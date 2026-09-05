"""Spreadsheets as a grid Mike can read and change.

`read_document` already turns a spreadsheet into prose, which is fine for
"what does this say" and useless for "put 4820 in C5". Cell work needs cell
addresses, so this is a second view of the same file rather than a second
kind of file: a grid in, a grid out, addressed the way every spreadsheet
application addresses it.

Two honesty constraints shape the whole module.

The first is that Mike has no calculation engine. openpyxl stores a formula
string; it does not evaluate it, and a file Mike wrote carries no cached
result for any formula in it. So a formula's value is reported as unknown
rather than guessed, and the caller is told to open the file in a real
spreadsheet application if it needs the number. Reporting a plausible total
that nothing computed is exactly the fabrication this runtime exists to
avoid.

The second is that a write is only reported as done after the file has been
read back off disk. openpyxl's save() succeeding says the library did not
raise; it does not say the cell holds what was asked.

Formats: .xlsx through openpyxl, .csv through the standard library. Apple's
.numbers is a proprietary bundle nothing here can write, so it is refused by
name with the export that would work.
"""
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from tools.filesystem.path_utils import resolve_path

# Enough for a real sheet, bounded so a 100k-row export cannot fill the
# model's context with data it will never read.
MAX_ROWS = 200
MAX_COLUMNS = 40

_CELL = re.compile(r"^([A-Za-z]{1,3})([1-9][0-9]{0,6})$")


class SpreadsheetError(Exception):
    """The file exists but cannot be handled as a spreadsheet."""


# ── addressing ────────────────────────────────────────────

def parse_cell(ref: str) -> tuple[int, int]:
    """'B3' -> (row 3, column 2). Raises on anything else."""
    match = _CELL.match((ref or "").strip())
    if not match:
        raise SpreadsheetError(
            f"{ref!r} is not a cell reference. Use a column letter followed by "
            "a row number, like B3."
        )
    letters, digits = match.group(1).upper(), match.group(2)
    column = 0
    for ch in letters:
        column = column * 26 + (ord(ch) - 64)
    return int(digits), column


def column_name(index: int) -> str:
    """1 -> 'A'. The inverse of the letter half of parse_cell."""
    name = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


# ── reading ───────────────────────────────────────────────

def _open_workbook(file: Path, values: bool):
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is installed
        raise SpreadsheetError(
            "openpyxl is not installed, so .xlsx files cannot be read."
        ) from exc
    try:
        return openpyxl.load_workbook(str(file), data_only=values)
    except Exception as exc:
        raise SpreadsheetError(f"{file.name} could not be opened as a workbook: {exc}") from exc


SUPPORTED = (".xlsx", ".xlsm", ".csv")


def _readable_siblings(file: Path) -> str:
    """What else is in the folder that Mike *can* open.

    An unsupported file is a dead end only if the agent has no idea what to
    try instead. Very often the export it is being told to make already
    exists a few bytes away, and the difference between "ask the user to
    export it" and "carry on with the export that is already there" is
    entirely whether the error mentioned it. The tool supplies the fact; the
    model still decides what to do with it.
    """
    try:
        others = sorted(
            p.name for p in file.parent.iterdir()
            if p.is_file() and p.suffix.lower() in SUPPORTED
        )
    except OSError:
        return ""
    if not others:
        return ""
    return (
        " These files in the same folder can be opened: "
        + ", ".join(others[:10])
        + ". If one of them is an export of this document, use it."
    )


def _check(file: Path) -> str:
    if not file.exists():
        raise FileNotFoundError(f"File not found: {file}")
    suffix = file.suffix.lower()
    if suffix == ".numbers":
        raise SpreadsheetError(
            f"{file.name} is an Apple Numbers document, which is a proprietary "
            "bundle Mike cannot read or write directly. Export it from Numbers "
            "as .xlsx or .csv (File > Export To) and work on that."
            + _readable_siblings(file)
        )
    if suffix not in SUPPORTED:
        raise SpreadsheetError(
            f"{file.name} is not a spreadsheet Mike can work with "
            f"({suffix or 'no extension'}). Supported: .xlsx, .xlsm, .csv."
            + _readable_siblings(file)
        )
    return suffix


def read_sheet(path: str, sheet: str | None = None) -> dict[str, Any]:
    """The grid, its formulas, and an explicit account of what is unknown."""
    file = resolve_path(path)
    suffix = _check(file)

    if suffix == ".csv":
        return _read_csv_grid(file)

    # Two passes on purpose: one workbook reports what each cell *contains*
    # (a formula string where there is one), the other reports the value the
    # authoring application last cached for it. Neither alone is the truth.
    formulas_wb = _open_workbook(file, values=False)
    values_wb = _open_workbook(file, values=True)

    names = formulas_wb.sheetnames
    if sheet:
        if sheet not in names:
            raise SpreadsheetError(
                f"{file.name} has no sheet called {sheet!r}. It has: {', '.join(names)}."
            )
        ws, vs = formulas_wb[sheet], values_wb[sheet]
    else:
        ws, vs = formulas_wb.active, values_wb[formulas_wb.active.title]

    rows: list[dict] = []
    formulas: dict[str, str] = {}
    uncomputed: list[str] = []
    truncated = False

    for r, (form_row, value_row) in enumerate(
        zip(ws.iter_rows(), vs.iter_rows()), start=1
    ):
        if r > MAX_ROWS:
            truncated = True
            break
        cells: dict[str, Any] = {}
        for c, (form_cell, value_cell) in enumerate(zip(form_row, value_row), start=1):
            if c > MAX_COLUMNS:
                truncated = True
                break
            raw, computed = form_cell.value, value_cell.value
            if raw is None and computed is None:
                continue
            ref = f"{column_name(c)}{r}"
            if isinstance(raw, str) and raw.startswith("="):
                formulas[ref] = raw
                if computed is None:
                    uncomputed.append(ref)
                    cells[column_name(c)] = f"{raw} (not calculated)"
                else:
                    cells[column_name(c)] = computed
            else:
                cells[column_name(c)] = raw if raw is not None else computed
        if cells:
            rows.append({"row": r, "cells": cells})

    formulas_wb.close()
    values_wb.close()

    result = {
        "status": "success",
        "path": str(file),
        "format": suffix.lstrip("."),
        "sheet": ws.title,
        "sheets": names,
        "rows": rows,
        "grid": render_grid(rows),
        "formulas": formulas,
    }
    if uncomputed:
        result["uncomputed_formulas"] = uncomputed
        result["note"] = (
            f"{len(uncomputed)} formula(s) have no calculated value stored in the "
            "file: " + ", ".join(uncomputed[:10]) + ". Mike does not evaluate "
            "formulas, so the number they produce is unknown here. Open the file "
            "in a spreadsheet application to see it, or write the arithmetic "
            "result as a literal if the number itself is what matters."
        )
    if truncated:
        result["truncated"] = (
            f"Showing at most {MAX_ROWS} rows and {MAX_COLUMNS} columns; the "
            "sheet is larger."
        )
    return result


def _read_csv_grid(file: Path) -> dict[str, Any]:
    text = file.read_text(encoding="utf-8", errors="replace")
    rows: list[dict] = []
    truncated = False
    for r, values in enumerate(csv.reader(text.splitlines()), start=1):
        if r > MAX_ROWS:
            truncated = True
            break
        cells = {
            column_name(c): v
            for c, v in enumerate(values[:MAX_COLUMNS], start=1)
            if v != ""
        }
        if cells:
            rows.append({"row": r, "cells": cells})
    result = {
        "status": "success",
        "path": str(file),
        "format": "csv",
        "sheet": file.stem,
        "sheets": [file.stem],
        "rows": rows,
        "grid": render_grid(rows),
        "formulas": {},
        "note": (
            "CSV stores text only. A cell beginning with '=' is stored as that "
            "literal text and no application will calculate it here."
        ),
    }
    if truncated:
        result["truncated"] = f"Showing the first {MAX_ROWS} rows."
    return result


def render_grid(rows: list[dict]) -> str:
    """A plain aligned table. Models read this far more reliably than JSON,
    and a person reading a log can see the sheet at a glance."""
    if not rows:
        return "(empty)"
    columns = sorted(
        {c for row in rows for c in row["cells"]},
        key=lambda name: parse_cell(f"{name}1")[1],
    )
    widths = {c: max(len(c), 1) for c in columns}
    for row in rows:
        for c in columns:
            widths[c] = max(widths[c], len(str(row["cells"].get(c, ""))))
    gutter = max(len(str(row["row"])) for row in rows)

    header = " " * gutter + " | " + " | ".join(c.ljust(widths[c]) for c in columns)
    lines = [header, "-" * len(header)]
    for row in rows:
        body = " | ".join(str(row["cells"].get(c, "")).ljust(widths[c]) for c in columns)
        lines.append(f"{str(row['row']).rjust(gutter)} | {body}")
    return "\n".join(lines)


# ── writing ───────────────────────────────────────────────

def _coerce(value: Any) -> Any:
    """Numbers typed as text are still numbers. A sheet full of strings that
    look like numbers is a sheet where SUM returns zero, so '4820' becomes
    4820 while '=SUM(B2:B5)' and 'Widget' are left exactly as given.

    An empty value becomes None, which is how a spreadsheet spells "this cell
    is empty". Writing "" instead leaves a cell that is blank to look at but
    present to every formula that counts cells.
    """
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    text = value.strip()
    if not text:
        return None
    if text.startswith("="):
        return text
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return value


def write_cells(
    path: str, cells: dict[str, Any], sheet: str | None = None
) -> dict[str, Any]:
    """Set cells and save, then read the file back and prove it took."""
    file = resolve_path(path)
    suffix = _check(file)

    if not cells:
        raise SpreadsheetError("No cells were given to write.")

    requested = {}
    for ref, value in cells.items():
        parse_cell(ref)  # raises with a usable message on a bad reference
        requested[ref.strip().upper()] = _coerce(value)

    if suffix == ".csv":
        target_sheet = _write_csv_cells(file, requested)
    else:
        target_sheet = _write_xlsx_cells(file, requested, sheet)

    # Independent verification: reopen from disk. Every mismatch is reported,
    # not the first one, so one call tells the model everything that is wrong.
    after = read_sheet(str(file), target_sheet)
    stored = {
        f"{column}{row['row']}": value
        for row in after["rows"]
        for column, value in row["cells"].items()
    }

    wrong = []
    for ref, wanted in requested.items():
        if isinstance(wanted, str) and wanted.startswith("=") and suffix != ".csv":
            # A formula is verified as a formula: the file must contain it.
            # Its value is a separate, unanswerable question here.
            if after["formulas"].get(ref) != wanted:
                wrong.append(f"{ref}: expected formula {wanted!r}, found "
                             f"{after['formulas'].get(ref)!r}")
            continue
        found = stored.get(ref)
        # A cleared cell reads back as absent, because that is what empty
        # means in a spreadsheet. Treating "asked for empty, found nothing"
        # as a mismatch reported a perfectly good write as a failure -- seen
        # in a real run, where it cost the agent a step to recover from a
        # problem that did not exist.
        if wanted is None:
            if found not in (None, ""):
                wrong.append(f"{ref}: expected it to be empty, found {found!r}")
            continue
        if str(found) != str(wanted):
            wrong.append(f"{ref}: expected {wanted!r}, found {found!r}")

    if wrong:
        return {
            "status": "error",
            "error": (
                "The file was saved but does not contain what was asked for: "
                + "; ".join(wrong)
            ),
            "path": str(file),
        }

    result = {
        "status": "success",
        "path": str(file),
        "sheet": target_sheet,
        "written": sorted(requested),
        "result": (
            f"Set {len(requested)} cell(s) in {file.name} "
            f"({', '.join(sorted(requested))}) and confirmed by reopening the file."
        ),
        "grid": after["grid"],
    }
    if after.get("uncomputed_formulas"):
        result["note"] = after["note"]
    return result


def _write_xlsx_cells(file: Path, cells: dict[str, Any], sheet: str | None) -> str:
    wb = _open_workbook(file, values=False)
    if sheet:
        if sheet not in wb.sheetnames:
            wb.close()
            raise SpreadsheetError(
                f"{file.name} has no sheet called {sheet!r}. It has: "
                f"{', '.join(wb.sheetnames)}."
            )
        ws = wb[sheet]
    else:
        ws = wb.active
    for ref, value in cells.items():
        ws[ref] = value
    try:
        wb.save(str(file))
    except PermissionError as exc:
        raise SpreadsheetError(
            f"{file.name} could not be saved: {exc}. It is most likely open in "
            "another application holding a lock on it."
        ) from exc
    finally:
        title = ws.title
        wb.close()
    return title


def _write_csv_cells(file: Path, cells: dict[str, Any]) -> str:
    text = file.read_text(encoding="utf-8", errors="replace")
    grid = [list(r) for r in csv.reader(text.splitlines())]

    for ref, value in cells.items():
        row, column = parse_cell(ref)
        while len(grid) < row:
            grid.append([])
        line = grid[row - 1]
        while len(line) < column:
            line.append("")
        line[column - 1] = str(value)

    width = max((len(r) for r in grid), default=0)
    for line in grid:
        line.extend([""] * (width - len(line)))

    with file.open("w", newline="", encoding="utf-8") as handle:
        csv.writer(handle).writerows(grid)
    return file.stem
