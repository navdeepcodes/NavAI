"""Cell-level spreadsheet work, and the two things it must never do:
claim a number it did not calculate, and report a save it did not verify.

Mike has no calculation engine. openpyxl stores a formula as a string and
never evaluates it, and a workbook Mike wrote carries no cached result for
any formula in it. The dangerous version of this capability is one that
writes '=SUM(C2:C5)' and then tells the user the total is 8000. So the read
path reports an uncalculated formula as unknown, out loud, and says where the
real number can be got.

The second guarantee is the one every other write in this runtime already
makes: openpyxl's save() returning means the library did not raise, not that
the cell holds what was asked. The file is reopened and checked.
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests import _isolate  # noqa: F401

import pytest

from tools.filesystem import spreadsheet as sheet_tools
from tools.filesystem.spreadsheet import SpreadsheetError

SALES = [
    ["Product", "Units", "Revenue"],
    ["Widget", 120, 2400],
    ["Gadget", 80, 3200],
    ["Doohickey", 45, 900],
]


@pytest.fixture
def workbook(tmp_path) -> str:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"
    for row in SALES:
        ws.append(row)
    path = tmp_path / "sales.xlsx"
    wb.save(str(path))
    return str(path)


@pytest.fixture
def sheet_csv(tmp_path) -> str:
    path = tmp_path / "sales.csv"
    path.write_text("\n".join(",".join(str(c) for c in row) for row in SALES) + "\n")
    return str(path)


# ── addressing ────────────────────────────────────────────

@pytest.mark.parametrize("ref,expected", [
    ("A1", (1, 1)), ("B3", (3, 2)), ("Z9", (9, 26)),
    ("AA12", (12, 27)), ("b3", (3, 2)),
])
def test_cell_references_parse(ref, expected):
    assert sheet_tools.parse_cell(ref) == expected


@pytest.mark.parametrize("bad", ["", "1A", "A", "3", "A0", "hello", "A1:B2"])
def test_bad_references_are_refused_with_a_usable_message(bad):
    """A model that mistypes a reference should be told how to type it, not
    handed an IndexError."""
    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.parse_cell(bad)
    assert "B3" in str(exc.value)


def test_column_names_round_trip():
    for index in (1, 2, 26, 27, 52, 53, 702, 703):
        assert sheet_tools.parse_cell(f"{sheet_tools.column_name(index)}1")[1] == index


# ── reading ───────────────────────────────────────────────

def test_reading_gives_addressed_cells(workbook):
    result = sheet_tools.read_sheet(workbook)

    assert result["status"] == "success"
    assert result["sheet"] == "Q3"
    assert result["sheets"] == ["Q3"]

    cells = {f"{c}{row['row']}": v for row in result["rows"] for c, v in row["cells"].items()}
    assert cells["A1"] == "Product"
    assert cells["B2"] == 120
    assert cells["C4"] == 900


def test_the_grid_is_readable(workbook):
    grid = sheet_tools.read_sheet(workbook)["grid"]
    assert "Product" in grid and "Doohickey" in grid
    assert "A" in grid.splitlines()[0] and "C" in grid.splitlines()[0]


def test_a_named_sheet_can_be_chosen(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "First"
    wb.active["A1"] = "one"
    wb.create_sheet("Second")["A1"] = "two"
    path = tmp_path / "two.xlsx"
    wb.save(str(path))

    assert sheet_tools.read_sheet(str(path))["sheet"] == "First"
    assert sheet_tools.read_sheet(str(path), "Second")["rows"][0]["cells"]["A"] == "two"


def test_a_missing_sheet_names_the_real_ones(workbook):
    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.read_sheet(workbook, "Q4")
    assert "Q3" in str(exc.value)


def test_csv_reads_as_a_grid(sheet_csv):
    result = sheet_tools.read_sheet(sheet_csv)
    assert result["format"] == "csv"
    assert result["rows"][1]["cells"]["A"] == "Widget"


# ── the formula honesty guarantee ─────────────────────────

def test_an_uncalculated_formula_is_reported_as_unknown(workbook):
    """The central guarantee. Mike wrote a formula; nothing evaluated it; the
    read must say so rather than leaving a number-shaped hole the model fills
    in from imagination."""
    sheet_tools.write_cells(workbook, {"C6": "=SUM(C2:C4)"})

    result = sheet_tools.read_sheet(workbook)

    assert result["formulas"]["C6"] == "=SUM(C2:C4)"
    assert "C6" in result["uncomputed_formulas"]
    assert "not calculated" in result["grid"]
    assert "does not evaluate formulas" in result["note"]


def test_a_cached_value_written_by_a_real_application_is_shown(tmp_path):
    """The other half: when the authoring application did compute a value,
    that value is reported. The rule is 'report what the file knows', not
    'never report formula values'."""
    from openpyxl import Workbook

    formulas = tmp_path / "cached.xlsx"
    wb = Workbook()
    wb.active["A1"] = 5
    wb.active["A2"] = "=A1*2"
    wb.save(str(formulas))

    # Stand in for the calculation a spreadsheet application would have done
    # and cached on save.
    import zipfile

    with zipfile.ZipFile(str(formulas)) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode()
    patched = sheet_xml.replace(
        '<f>A1*2</f>', '<f>A1*2</f><v>10</v>'
    )
    rebuilt = tmp_path / "rebuilt.xlsx"
    with zipfile.ZipFile(str(formulas)) as src, zipfile.ZipFile(str(rebuilt), "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = patched.encode()
            dst.writestr(item, data)

    result = sheet_tools.read_sheet(str(rebuilt))
    cells = {f"{c}{r['row']}": v for r in result["rows"] for c, v in r["cells"].items()}
    assert cells["A2"] == 10
    assert "uncomputed_formulas" not in result


def test_csv_says_formulas_are_only_text(sheet_csv):
    result = sheet_tools.read_sheet(sheet_csv)
    assert "no application will calculate it" in result["note"]


# ── writing ───────────────────────────────────────────────

def test_writing_cells_persists_to_disk(workbook):
    result = sheet_tools.write_cells(
        workbook, {"A5": "Gizmo", "B5": 60, "C5": 1500}
    )
    assert result["status"] == "success"
    assert result["written"] == ["A5", "B5", "C5"]

    # Read with a fresh call, not from the write's own return value.
    reopened = sheet_tools.read_sheet(workbook)
    cells = {f"{c}{r['row']}": v for r in reopened["rows"] for c, v in r["cells"].items()}
    assert cells["A5"] == "Gizmo"
    assert cells["C5"] == 1500


def test_numeric_text_is_stored_as_a_number(workbook):
    """A model producing JSON hands over "60" far more often than 60. Stored
    as text, it makes every SUM below it silently wrong, which is the worst
    kind of wrong — it looks fine."""
    sheet_tools.write_cells(workbook, {"B5": "60", "C5": "1500.5"})

    from openpyxl import load_workbook

    ws = load_workbook(workbook).active
    assert ws["B5"].value == 60 and isinstance(ws["B5"].value, int)
    assert ws["C5"].value == 1500.5


def test_a_word_that_is_not_a_number_stays_text(workbook):
    sheet_tools.write_cells(workbook, {"A5": "Gizmo"})
    from openpyxl import load_workbook
    assert load_workbook(workbook).active["A5"].value == "Gizmo"


def test_a_formula_is_stored_as_a_formula(workbook):
    sheet_tools.write_cells(workbook, {"C6": "=SUM(C2:C5)"})
    from openpyxl import load_workbook
    assert load_workbook(workbook).active["C6"].value == "=SUM(C2:C5)"


def test_writing_does_not_disturb_other_cells(workbook):
    sheet_tools.write_cells(workbook, {"B5": 60})
    reopened = sheet_tools.read_sheet(workbook)
    cells = {f"{c}{r['row']}": v for r in reopened["rows"] for c, v in r["cells"].items()}
    for ref, expected in (("A1", "Product"), ("B2", 120), ("C4", 900)):
        assert cells[ref] == expected, f"{ref} changed"


def test_csv_writing_persists(sheet_csv):
    sheet_tools.write_cells(sheet_csv, {"A5": "Gizmo", "B5": 60})
    text = Path(sheet_csv).read_text()
    assert "Gizmo,60" in text
    assert "Widget,120,2400" in text


def test_csv_write_extends_a_short_row(sheet_csv):
    sheet_tools.write_cells(sheet_csv, {"E9": "far out"})
    reopened = sheet_tools.read_sheet(sheet_csv)
    cells = {f"{c}{r['row']}": v for r in reopened["rows"] for c, v in r["cells"].items()}
    assert cells["E9"] == "far out"


# ── verification and failure ──────────────────────────────

def test_a_save_that_did_not_take_is_reported_as_failure(workbook, monkeypatch):
    """The verification must be real. If the file silently keeps its old
    contents, the tool has to say so instead of reporting success — this
    simulates exactly that."""
    import openpyxl

    monkeypatch.setattr(openpyxl.Workbook, "save", lambda self, path: None)

    result = sheet_tools.write_cells(workbook, {"B5": 60})

    assert result["status"] == "error"
    assert "B5" in result["error"]
    assert "does not contain what was asked for" in result["error"]


def test_a_missing_file_is_a_missing_file(tmp_path):
    with pytest.raises(FileNotFoundError):
        sheet_tools.read_sheet(str(tmp_path / "nope.xlsx"))


def test_a_numbers_document_says_how_to_convert_it(tmp_path):
    """The one spreadsheet application on this Mac writes a format nothing
    here can parse. Refusing by name with the export that works is more
    useful than a parse error."""
    path = tmp_path / "sales.numbers"
    path.write_bytes(b"PK\x03\x04 not really")
    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.read_sheet(str(path))
    assert "Export it from Numbers" in str(exc.value)
    assert ".xlsx" in str(exc.value)


def test_an_unsupported_extension_lists_what_works(tmp_path):
    path = tmp_path / "notes.txt"
    path.write_text("hello")
    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.read_sheet(str(path))
    assert ".xlsx" in str(exc.value) and ".csv" in str(exc.value)


def test_a_corrupt_workbook_is_reported_not_crashed(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"this is not a zip archive at all")
    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.read_sheet(str(path))
    assert "broken.xlsx" in str(exc.value)


def test_writing_nothing_is_refused(workbook):
    with pytest.raises(SpreadsheetError):
        sheet_tools.write_cells(workbook, {})


# ── the runtime surface ───────────────────────────────────

def test_the_tools_are_declared_to_the_model():
    from brain.core_tools import TOOL_DECLARATIONS

    names = {d.name for d in TOOL_DECLARATIONS}
    assert {"read_spreadsheet", "edit_spreadsheet"} <= names


def test_changing_cells_requires_confirmation_and_reading_does_not():
    from brain.core_tools import needs_confirmation

    assert needs_confirmation("edit_spreadsheet", {"path": "x.xlsx", "cells": {}})
    assert not needs_confirmation("read_spreadsheet", {"path": "x.xlsx"})


def test_the_confirmation_shows_the_values_being_overwritten(workbook):
    """Read from the file, not from the arguments: the user needs to see that
    C4 currently holds 900 before agreeing to replace it."""
    from brain.core_tools import describe_action

    prompt = describe_action(
        "edit_spreadsheet", {"path": workbook, "cells": {"C4": 1000, "A5": "Gizmo"}}
    )
    assert "900" in prompt, prompt
    assert "1000" in prompt
    assert "(empty)" in prompt, "a new cell should be shown as empty, not omitted"


def test_the_runtime_returns_the_grid_intact(workbook):
    """Structured evidence must reach the model, not a flattened sentence."""
    from brain.core_runtime import CoreRuntime

    result = CoreRuntime()._execute_tool("read_spreadsheet", {"path": workbook})
    assert result["status"] == "success"
    assert "grid" in result and "Doohickey" in result["grid"]


def test_the_runtime_explains_a_wrongly_typed_cells_argument(workbook):
    from brain.core_runtime import CoreRuntime

    result = CoreRuntime()._execute_tool(
        "edit_spreadsheet", {"path": workbook, "cells": "B5=60"}
    )
    assert result["status"] == "error"
    assert result.get("retry_safe") is True
    assert "object" in result["error"]


def test_the_runtime_keeps_a_missing_file_retryable(tmp_path):
    from brain.core_runtime import CoreRuntime

    result = CoreRuntime()._execute_tool(
        "read_spreadsheet", {"path": str(tmp_path / "nope.xlsx")}
    )
    assert result["status"] == "error"
    assert result["retry_safe"] is True


def test_an_unsupported_file_names_what_can_be_opened(tmp_path):
    """An error that only says "I can't read this" ends the task. Naming the
    readable files beside it gives the agent something to act on, and it is
    the agent that decides whether one of them is the right file.

    This came out of a real run: given a .numbers document with an .xlsx
    export sitting next to it, Mike correctly refused the .numbers, correctly
    did not invent figures — and then stopped to ask the user to export a
    file that already existed, because nothing had told it so."""
    (tmp_path / "Q3 Sales.numbers").write_bytes(b"PK\x03\x04not-a-bundle")
    (tmp_path / "q3_sales.xlsx").write_bytes(b"placeholder")
    (tmp_path / "notes.txt").write_text("ignore me")

    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.read_sheet(str(tmp_path / "Q3 Sales.numbers"))

    message = str(exc.value)
    assert "q3_sales.xlsx" in message
    assert "notes.txt" not in message, "it offered a file it cannot open either"


def test_the_sibling_hint_is_omitted_when_there_is_nothing_to_offer(tmp_path):
    (tmp_path / "Q3 Sales.numbers").write_bytes(b"PK\x03\x04not-a-bundle")
    with pytest.raises(SpreadsheetError) as exc:
        sheet_tools.read_sheet(str(tmp_path / "Q3 Sales.numbers"))
    assert "same folder" not in str(exc.value)


def test_clearing_a_cell_is_not_reported_as_a_failed_write(workbook):
    """From a real cross-application run: the agent wrote "" into a cell to
    leave it blank, which is right, and the verification then told it the
    write had failed because the cell read back as absent. Empty is absent —
    that is what a spreadsheet means by empty — so the two must agree."""
    result = sheet_tools.write_cells(workbook, {"A5": "Total", "B5": "", "C5": 9304})

    assert result["status"] == "success", result.get("error")

    from openpyxl import load_workbook
    ws = load_workbook(workbook).active
    assert ws["B5"].value is None, "an empty value should clear the cell, not store ''"
    assert ws["C5"].value == 9304


def test_clearing_an_existing_cell_actually_clears_it(workbook):
    sheet_tools.write_cells(workbook, {"B2": ""})
    from openpyxl import load_workbook
    assert load_workbook(workbook).active["B2"].value is None
