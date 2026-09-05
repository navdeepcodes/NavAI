"""Real agent test: inspect a spreadsheet, add data, total it, save, verify.

The agent is given a goal and the whole tool set. Nothing sequences the work
and nothing tells it that read_spreadsheet or edit_spreadsheet exist beyond
their ordinary declarations.

Two properties make the result trustworthy.

**Verification reads the file.** Every check below reopens the workbook with
openpyxl and looks at cells. The model's account of what it did is printed
but never used to decide whether the run passed.

**The total has to be right, not merely present.** Mike cannot evaluate
formulas, so a run that writes =SUM(...) and stops has produced a file whose
total is unknown. The task asks for the number; the check requires the stored
value to equal the arithmetic. A formula alongside it is welcome and checked
separately.

    venv/bin/python tests/endurance_spreadsheet.py [--steps N] [--case recovery]
"""
from __future__ import annotations

import argparse
import os
import shutil
import sys
import time
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests.harness import report, run_agent, save

WORKSPACE = Path.home() / "mike-spreadsheet-task"

# Deliberately not round numbers: a plausible-looking total cannot be guessed.
SALES = [
    ("Product", "Units", "Revenue"),
    ("Widget", 120, 2417),
    ("Gadget", 83, 3168),
    ("Doohickey", 46, 912),
    ("Sprocket", 71, 1543),
]
NEW_ROW = ("Flywheel", 58, 1264)
EXPECTED_TOTAL = sum(row[2] for row in SALES[1:]) + NEW_ROW[2]  # 9304


def build(case: str) -> Path:
    if WORKSPACE.exists():
        shutil.rmtree(WORKSPACE)
    WORKSPACE.mkdir(parents=True)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"
    for row in SALES:
        ws.append(list(row))
    path = WORKSPACE / "q3_sales.xlsx"
    wb.save(str(path))

    if case in ("recovery", "numbers"):
        # A decoy in the format Mike genuinely cannot parse, named the way a
        # user would name the file they mean. The agent has to read the error,
        # understand that .numbers is not workable, look in the folder, and
        # carry on with the .xlsx — without inventing figures from the name.
        (WORKSPACE / "Q3 Sales.numbers").write_bytes(b"PK\x03\x04not-a-real-bundle")
    return path


TASK = """In the folder {folder} there is a spreadsheet of Q3 sales.

Open it and look at the figures. Then add a new product row for {product}: \
{units} units, revenue {revenue}.

After that, add a Total row underneath with the total revenue for all \
products, including the new one. I want to see the actual number in the \
sheet, not just a formula.

Save the file, then check it really contains what you say it does."""

NUMBERS_TASK = """Open {folder}/"Q3 Sales.numbers" and tell me the total \
revenue across all the products in it.

Then add a Total row at the bottom of the sheet with that number in the \
revenue column, and save it."""

RECOVERY_TASK = """Open the Q3 sales spreadsheet in {folder} — I think it's \
called "Q3 Sales" — and tell me the total revenue across all the products \
listed in it.

Then add a Total row at the bottom of the sheet with that number in the \
revenue column, and save it."""


def verify(path: Path, case: str) -> list[tuple[str, bool, str]]:
    """Objective checks. Nothing here reads the model's reply."""
    from openpyxl import load_workbook

    checks: list[tuple[str, bool, str]] = []

    if not path.exists():
        return [("the spreadsheet still exists", False, f"{path} is gone")]

    try:
        ws = load_workbook(str(path)).active
        formula_ws = load_workbook(str(path)).active
    except Exception as exc:
        return [("the file is still a valid workbook", False, str(exc))]

    grid = {
        f"{cell.column_letter}{cell.row}": cell.value
        for row in ws.iter_rows() for cell in row if cell.value is not None
    }

    checks.append((
        "the original data is untouched",
        grid.get("A2") == "Widget" and grid.get("C3") == 3168,
        f"A2={grid.get('A2')!r} C3={grid.get('C3')!r}",
    ))

    if case == "normal":
        added = [ref for ref, value in grid.items()
                 if isinstance(value, str) and value.strip().lower() == NEW_ROW[0].lower()]
        checks.append((
            f"the {NEW_ROW[0]} row was added",
            bool(added),
            f"found at {added}" if added else f"no {NEW_ROW[0]} anywhere in the sheet",
        ))
        if added:
            row = int(added[0][1:])
            checks.append((
                "its units and revenue are right",
                grid.get(f"B{row}") == NEW_ROW[1] and grid.get(f"C{row}") == NEW_ROW[2],
                f"B{row}={grid.get(f'B{row}')!r} C{row}={grid.get(f'C{row}')!r} "
                f"(wanted {NEW_ROW[1]}, {NEW_ROW[2]})",
            ))

    wanted_total = EXPECTED_TOTAL if case == "normal" else sum(r[2] for r in SALES[1:])

    total_cells = [
        ref for ref, value in grid.items()
        if ref.startswith("C") and isinstance(value, (int, float)) and value == wanted_total
    ]
    checks.append((
        f"the revenue total {wanted_total} is stored as a number",
        bool(total_cells),
        f"found in {total_cells}" if total_cells else
        f"no cell in column C holds {wanted_total}; C column holds "
        + repr({k: v for k, v in grid.items() if k.startswith("C")}),
    ))

    labelled = [
        ref for ref, value in grid.items()
        if isinstance(value, str) and "total" in value.lower()
    ]
    checks.append((
        "the total row is labelled",
        bool(labelled),
        f"label at {labelled}" if labelled else "nothing in the sheet says Total",
    ))

    # Not required, but recorded: a formula alongside the number is the ideal
    # outcome and worth knowing about.
    formulas = {
        f"{c.column_letter}{c.row}": c.value
        for row in formula_ws.iter_rows() for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    }
    print(f"\n  (formulas present in the saved file: {formulas or 'none'})")

    return checks


def _same_file(given, expected: Path) -> bool:
    """Is this the same file on disk?

    Not a string comparison. macOS filesystems are case-insensitive, so a
    model that writes Q3_sales.xlsx opens exactly the file that q3_sales.xlsx
    names — and an earlier version of this check compared resolved paths as
    text and refused a perfectly correct edit, failing a run in which Mike had
    done everything right. samefile asks the filesystem instead of guessing.
    """
    if not given:
        return False
    candidate = Path(os.path.expanduser(str(given)))
    try:
        return candidate.samefile(expected)
    except OSError:
        return False


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--steps", type=int, default=None)
    parser.add_argument("--case", default="normal",
                        choices=["normal", "recovery", "numbers"])
    parser.add_argument("--run", type=int, default=1)
    args = parser.parse_args()

    path = build(args.case)

    if args.case == "numbers":
        task = NUMBERS_TASK.format(folder=WORKSPACE)
    elif args.case == "recovery":
        task = RECOVERY_TASK.format(folder=WORKSPACE)
    else:
        task = TASK.format(
            folder=WORKSPACE, product=NEW_ROW[0],
            units=NEW_ROW[1], revenue=NEW_ROW[2],
        )

    print(f"workspace : {WORKSPACE}")
    print(f"case      : {args.case}")
    print(f"expected  : total {EXPECTED_TOTAL if args.case == 'normal' else sum(r[2] for r in SALES[1:])}\n")

    seen: list[str] = []

    def confirm(detail, tool, tool_args):
        """Stand in for the user. Only spreadsheet edits to the intended file
        are allowed — a run that tries to write somewhere else is refused
        here rather than passing quietly."""
        seen.append(tool)
        if tool != "edit_spreadsheet":
            return False
        return _same_file(tool_args.get("path"), path)

    run = run_agent(task, confirm=confirm, steps=args.steps, time_budget=1800)
    run.checks = verify(path, args.case)
    run.checks.append((
        "it asked before changing the file",
        any(c["tool"] == "edit_spreadsheet" for c in run.confirmations),
        f"confirmations: {[c['tool'] for c in run.confirmations] or 'none'}",
    ))

    report(run, f"spreadsheet ({args.case})")
    save(run, f"spreadsheet_{args.case}_run{args.run}")


if __name__ == "__main__":
    main()
