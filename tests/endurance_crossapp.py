"""Real agent test: a plain file, a spreadsheet, and Gmail, in one task.

This is the question the individual capability tests cannot answer — whether
Mike's tools compose. Nothing here sequences the work. The agent is given a
goal and the ordinary tool set, and it has to decide for itself to read the
notes, put the figures in the workbook, do the arithmetic, save, attach the
file it just changed, and send it.

Verification crosses every boundary the task does:

  * the workbook is reopened from disk and its cells are read
  * the confirmation for the send is judged against the real arguments, and
    the attachment is checked to be the updated workbook before allowing it
  * the message is fetched back from Gmail, and its attachment is downloaded
    and opened as a workbook — so the final check is that the bytes Google
    stored contain the total Mike calculated

A run that emails a stale copy, or the wrong file, or the right file with the
wrong number in it, fails at the last check rather than passing on the
strength of the model's summary.

    venv/bin/python tests/endurance_crossapp.py [--steps N] [--run N]
"""
from __future__ import annotations

import argparse
import base64
import os
import shutil
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests.harness import report, run_agent, save

WORKSPACE = Path.home() / "mike-crossapp-task"
WORKBOOK = WORKSPACE / "q3_sales.xlsx"
NOTES = WORKSPACE / "october_figures.txt"
def _recipient() -> str:
    """Where a real run sends its mail.

    Read from the environment rather than written down here. This repository
    is public, and a person's mailbox is not part of the test — anyone
    running this should point it at their own:

        MIKE_TEST_RECIPIENT=you@example.com venv/bin/python tests/endurance_crossapp.py
    """
    address = os.environ.get("MIKE_TEST_RECIPIENT", "").strip()
    if not address:
        raise SystemExit(
            "Set MIKE_TEST_RECIPIENT to the address this run should mail, "
            "e.g. MIKE_TEST_RECIPIENT=you@example.com"
        )
    return address


RECIPIENT = _recipient()
SUBJECT = "Q3 Sales Updated"

EXISTING = [
    ("Product", "Units", "Revenue"),
    ("Widget", 120, 2417),
    ("Gadget", 83, 3168),
    ("Doohickey", 46, 912),
]
# Only in the notes file, so the figures cannot come from the workbook alone.
OCTOBER = [("Sprocket", 71, 1543), ("Flywheel", 58, 1264)]
EXPECTED_TOTAL = sum(r[2] for r in EXISTING[1:]) + sum(r[2] for r in OCTOBER)  # 9304


def build() -> None:
    if WORKSPACE.exists():
        shutil.rmtree(WORKSPACE)
    WORKSPACE.mkdir(parents=True)

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Q3"
    for row in EXISTING:
        ws.append(list(row))
    wb.save(str(WORKBOOK))

    NOTES.write_text(
        "October figures (came in after the sheet was made)\n"
        "--------------------------------------------------\n"
        + "\n".join(f"{name}: {units} units, revenue {revenue}"
                    for name, units, revenue in OCTOBER)
        + "\n"
    )


TASK = f"""In {WORKSPACE} there's a sales spreadsheet and a text file with \
some figures that came in late.

Add the late figures to the spreadsheet as new product rows, then put a \
Total row at the bottom with the total revenue for everything in the sheet. \
I want the actual number there.

Save it, then email me the updated spreadsheet at {RECIPIENT} with the \
subject "{SUBJECT}". In the body, tell me what the total came to. Attach the \
spreadsheet file itself."""


def read_workbook(path: Path) -> dict:
    from openpyxl import load_workbook

    ws = load_workbook(str(path)).active
    return {
        f"{cell.column_letter}{cell.row}": cell.value
        for row in ws.iter_rows() for cell in row if cell.value is not None
    }


def _same_file(given, expected: Path) -> bool:
    """Is this the same file on disk?

    Not a string comparison. macOS filesystems are case-insensitive, so a
    model that writes Q3_sales.xlsx opens exactly the file that q3_sales.xlsx
    names — and an earlier version of this check compared resolved paths as
    text and refused a perfectly correct edit, failing a run in which Mike had
    done everything right. samefile asks the filesystem instead of guessing.
    """
    if not given:
        return False
    candidate = Path(os.path.expanduser(str(given)))
    try:
        return candidate.samefile(expected)
    except OSError:
        return False


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--steps", type=int, default=None)
    parser.add_argument("--run", type=int, default=1)
    args = parser.parse_args()

    build()
    print(f"workspace : {WORKSPACE}")
    print(f"recipient : {RECIPIENT}")
    print(f"expected  : total {EXPECTED_TOTAL}\n")

    transitions: list[str] = []
    send_args: dict = {}

    def confirm(detail, tool, tool_args):
        """Stand in for the user. This is not a rubber stamp: the send is
        allowed only if the recipient is right and the attachment on disk is
        the workbook, already carrying the total."""
        if tool == "edit_spreadsheet":
            return _same_file(tool_args.get("path"), WORKBOOK)

        if tool == "send_email":
            send_args.clear()
            send_args.update(tool_args)
            to = str(tool_args.get("to") or "").strip().lower()
            if to != RECIPIENT.lower():
                print(f"     (refused: recipient is {to!r})")
                return False
            attachments = [Path(os.path.expanduser(str(a)))
                           for a in (tool_args.get("attachments") or [])]
            for path in attachments:
                if path.suffix.lower() != ".xlsx" or not path.is_file():
                    continue
                try:
                    cells = read_workbook(path)
                except Exception as exc:
                    print(f"     (refused: {path.name} is not a workbook: {exc})")
                    return False
                if EXPECTED_TOTAL in [v for v in cells.values() if isinstance(v, (int, float))]:
                    return True
                print(f"     (refused: {path.name} does not contain {EXPECTED_TOTAL})")
                return False
            print("     (refused: no spreadsheet attached)")
            return False

        return False

    run = run_agent(TASK, confirm=confirm, steps=args.steps, time_budget=2400)

    # Which capability each call belonged to, in order — the record of the
    # transitions the task actually made.
    families = {
        "read_file": "filesystem", "read_document": "filesystem",
        "list_directory": "filesystem", "search_files": "filesystem",
        "read_lines": "filesystem", "write_file": "filesystem",
        "read_spreadsheet": "spreadsheet", "edit_spreadsheet": "spreadsheet",
        "send_email": "gmail",
    }
    for call in run.calls:
        family = families.get(call.tool, call.tool)
        if not transitions or transitions[-1] != family:
            transitions.append(family)
    print(f"\napplication transitions: {' -> '.join(transitions)}")

    # ── verification ──────────────────────────────────────
    checks: list[tuple[str, bool, str]] = []
    cells = read_workbook(WORKBOOK) if WORKBOOK.exists() else {}

    checks.append((
        "the original rows survived",
        cells.get("A2") == "Widget" and cells.get("C3") == 3168,
        f"A2={cells.get('A2')!r} C3={cells.get('C3')!r}",
    ))

    for name, units, revenue in OCTOBER:
        found = [ref for ref, v in cells.items()
                 if isinstance(v, str) and v.strip().lower() == name.lower()]
        ok = False
        detail = f"{name} is not in the sheet"
        if found:
            row = int(found[0][1:])
            ok = cells.get(f"C{row}") == revenue
            detail = f"row {row}: units={cells.get(f'B{row}')!r} revenue={cells.get(f'C{row}')!r}"
        checks.append((f"{name} was carried over from the notes file", ok, detail))

    total_at = [ref for ref, v in cells.items()
                if ref.startswith("C") and v == EXPECTED_TOTAL]
    checks.append((
        f"the total {EXPECTED_TOTAL} is in the sheet as a number",
        bool(total_at),
        f"at {total_at}" if total_at else
        "column C holds " + repr({k: v for k, v in cells.items() if k.startswith("C")}),
    ))

    checks.append((
        "it asked before changing the file",
        any(c["tool"] == "edit_spreadsheet" for c in run.confirmations),
        f"confirmed: {[c['tool'] for c in run.confirmations] or 'nothing'}",
    ))
    send_confirms = [c for c in run.confirmations if c["tool"] == "send_email"]
    checks.append((
        "it asked before sending",
        bool(send_confirms),
        send_confirms[-1]["detail"][:200] if send_confirms else "never asked",
    ))
    checks.append((
        "the send was allowed only with the right file attached",
        bool(send_confirms) and send_confirms[-1]["allowed"],
        "allowed" if send_confirms and send_confirms[-1]["allowed"] else
        "the stand-in refused every send it was shown",
    ))

    # ── the mailbox, and the bytes Google stored ──────────
    verified: dict = {}
    attachment_total = None
    try:
        from tools.email.gmail_client import GmailClient

        client = GmailClient()
        listing = client.service.users().messages().list(
            userId="me", q=f'subject:"{SUBJECT}" in:sent', maxResults=1,
        ).execute()
        ids = [m["id"] for m in listing.get("messages", [])]
        if ids:
            verified = client.describe_sent(ids[0])

            message = client.service.users().messages().get(
                userId="me", id=ids[0], format="full",
            ).execute()

            def find_part(part):
                if part.get("filename", "").lower().endswith(".xlsx"):
                    return part
                for child in part.get("parts", []) or []:
                    hit = find_part(child)
                    if hit:
                        return hit
                return None

            part = find_part(message.get("payload", {}))
            if part:
                body = part["body"]
                data = body.get("data")
                if not data and body.get("attachmentId"):
                    data = client.service.users().messages().attachments().get(
                        userId="me", messageId=ids[0], id=body["attachmentId"],
                    ).execute()["data"]
                raw = base64.urlsafe_b64decode(data)
                downloaded = Path(tempfile.mkdtemp()) / part["filename"]
                downloaded.write_bytes(raw)
                sent_cells = read_workbook(downloaded)
                attachment_total = [
                    ref for ref, v in sent_cells.items() if v == EXPECTED_TOTAL
                ]
    except Exception as exc:
        verified = {"error": f"{type(exc).__name__}: {exc}"}

    checks.append((
        "the message is in Sent",
        bool(verified.get("in_sent")),
        str(verified)[:220],
    ))
    checks.append((
        "the mailbox shows the right recipient and subject",
        RECIPIENT in str(verified.get("to", "")) and SUBJECT in str(verified.get("subject", "")),
        f"to={verified.get('to')!r} subject={verified.get('subject')!r}",
    ))
    checks.append((
        "an .xlsx is attached to the sent message",
        any(str(f).lower().endswith(".xlsx") for f in verified.get("attachments", [])),
        f"attachments: {verified.get('attachments')}",
    ))
    checks.append((
        f"the attachment Google stored contains the total {EXPECTED_TOTAL}",
        bool(attachment_total),
        f"found at {attachment_total}" if attachment_total else
        "the downloaded attachment does not contain the total",
    ))

    run.checks = checks
    report(run, "cross-application")
    print(f"transitions: {' -> '.join(transitions)}")
    save(run, f"crossapp_run{args.run}")


if __name__ == "__main__":
    main()
