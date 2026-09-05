"""Characterising what happens to Mike over a long task.

The benchmark tasks are five to seven steps. This one cannot be done in five
to seven: six workbooks each have to be read, totalled and written, and then
a summary has to be produced that depends on figures gathered at the very
beginning. It is built so that the failure modes worth knowing about have
somewhere to show themselves —

  * **context growth** — every turn's estimated request size is recorded,
    along with the budget it was measured against
  * **history trimming** — how many messages the planner dropped, and when
  * **state loss** — the summary needs the total from workbook one, which is
    the oldest thing in the conversation by the time it is needed
  * **forgotten goal** — the final report is checked against the whole task,
    not the last instruction in it
  * **repetition** — identical tool calls are counted, since re-reading a
    file it already read is what a model does when it has lost the answer
  * **the step limit** — whether the run ends because the work is done or
    because MAX_AGENT_STEPS was reached, and what Mike says when it is

Nothing here is a pass/fail benchmark for the model. It is a measurement of
the runtime under length, so the question "is the existing architecture
enough?" can be answered with numbers instead of an opinion.

    venv/bin/python tests/endurance_longhorizon.py [--steps N] [--files N]
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from tests.harness import report, run_agent, save

WORKSPACE = Path.home() / "mike-longhorizon-task"

REGIONS = [
    ("north", [("Widget", 2417), ("Gadget", 3168), ("Doohickey", 912)]),
    ("south", [("Widget", 1841), ("Gadget", 2260), ("Sprocket", 1543)]),
    ("east", [("Gadget", 3097), ("Flywheel", 1264), ("Doohickey", 806)]),
    ("west", [("Widget", 2255), ("Sprocket", 1719), ("Flywheel", 934)]),
    ("central", [("Widget", 1408), ("Gadget", 2871), ("Doohickey", 1122)]),
    ("islands", [("Sprocket", 655), ("Flywheel", 1490), ("Widget", 733)]),
]


def build(count: int) -> dict[str, int]:
    if WORKSPACE.exists():
        shutil.rmtree(WORKSPACE)
    WORKSPACE.mkdir(parents=True)

    from openpyxl import Workbook

    totals = {}
    for name, rows in REGIONS[:count]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Product", "Revenue"])
        for product, revenue in rows:
            ws.append([product, revenue])
        wb.save(str(WORKSPACE / f"{name}.xlsx"))
        totals[name] = sum(r[1] for r in rows)
    return totals


TASK = """There are several regional sales spreadsheets in {folder}.

For each one: open it, work out the total revenue, and add a Total row at the \
bottom of that sheet with the number in the revenue column.

When you've done all of them, create a file called summary.csv in the same \
folder with two columns, region and total, listing every region and the total \
you calculated for it, plus a final row for the combined total across all \
regions.

Work through them one at a time and don't skip any."""


def verify(totals: dict[str, int]) -> list[tuple[str, bool, str]]:
    from openpyxl import load_workbook

    checks: list[tuple[str, bool, str]] = []
    written = {}

    for region, expected in totals.items():
        path = WORKSPACE / f"{region}.xlsx"
        try:
            ws = load_workbook(str(path)).active
            values = [c.value for row in ws.iter_rows() for c in row]
        except Exception as exc:
            checks.append((f"{region}.xlsx has its total", False, str(exc)))
            continue
        ok = expected in values
        written[region] = ok
        checks.append((
            f"{region}.xlsx has a Total row of {expected}",
            ok,
            f"values in the sheet: {[v for v in values if isinstance(v, (int, float))]}",
        ))

    summary = WORKSPACE / "summary.csv"
    if not summary.exists():
        checks.append(("summary.csv was created", False, "the file does not exist"))
        return checks

    text = summary.read_text()
    checks.append(("summary.csv was created", True, text.strip()[:300]))

    # The point of the whole exercise: are the earliest figures still right by
    # the time the summary is written?
    for region, expected in totals.items():
        checks.append((
            f"the summary carries {region}'s real total ({expected})",
            str(expected) in text,
            f"summary text: {text.strip()[:200]}",
        ))

    combined = sum(totals.values())
    checks.append((
        f"the summary has the combined total ({combined})",
        str(combined) in text,
        f"summary text: {text.strip()[:200]}",
    ))
    return checks


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--steps", type=int, default=None)
    parser.add_argument("--files", type=int, default=6)
    parser.add_argument("--run", type=int, default=1)
    args = parser.parse_args()

    totals = build(args.files)
    print(f"workspace : {WORKSPACE}")
    print(f"workbooks : {len(totals)} — {totals}")

    # ── measure the planner, turn by turn ─────────────────
    import brain.context_budget as context_budget

    planning: list[dict] = []
    original_plan = context_budget.plan_request

    def measuring(messages, tools, capabilities, **kwargs):
        plan = original_plan(messages, tools, capabilities, **kwargs)
        planning.append({
            "messages_in": len(messages),
            "messages_sent": len(plan.messages),
            "tokens": plan.estimated_tokens,
            "budget": plan.budget,
            "dropped_history": plan.dropped_history,
            "dropped_tools": plan.dropped_tools,
            "notes": list(plan.notes),
            "error": plan.error.message if plan.error else "",
        })
        return plan

    # core_runtime imported the name directly, so patching the module it came
    # from would not be observed.
    import brain.core_runtime as core_runtime
    context_budget.plan_request = measuring
    core_runtime.plan_request = measuring

    def confirm(detail, tool, tool_args):
        """Allow the ordinary work of the task inside its own folder, and
        nothing else."""
        path = str(tool_args.get("path") or "")
        resolved = Path(os.path.expanduser(path)).resolve() if path else None
        inside = bool(resolved and WORKSPACE.resolve() in resolved.parents)
        return tool in ("edit_spreadsheet", "write_file", "create_file") and inside

    try:
        run = run_agent(TASK.format(folder=WORKSPACE), confirm=confirm,
                        steps=args.steps, time_budget=3000)
    finally:
        context_budget.plan_request = original_plan
        core_runtime.plan_request = original_plan

    run.checks = verify(totals)
    report(run, "long-horizon")

    # ── the characterisation ──────────────────────────────
    seen: dict[tuple, int] = {}
    for call in run.calls:
        key = (call.tool, json.dumps(call.args, sort_keys=True))
        seen[key] = seen.get(key, 0) + 1
    repeats = {f"{k[0]} {k[1][:80]}": n for k, n in seen.items() if n > 1}

    tokens = [p["tokens"] for p in planning] or [0]
    dropped = sum(p["dropped_history"] for p in planning)

    print("\n--- long-horizon characterisation ---")
    print(f"  turns                  : {run.turns}")
    print(f"  tool calls             : {len(run.calls)}")
    print(f"  step limit             : {args.steps or 'default'}"
          f"{' (REACHED)' if run.stopped_early else ''}")
    print(f"  request size, first    : {tokens[0]:,} tokens")
    print(f"  request size, largest  : {max(tokens):,} tokens")
    print(f"  request size, last     : {tokens[-1]:,} tokens")
    print(f"  budget                 : {planning[0]['budget']:,} tokens"
          if planning else "  budget: unknown")
    print(f"  peak use of budget     : "
          f"{max(tokens) / planning[0]['budget']:.0%}" if planning else "")
    print(f"  history messages dropped: {dropped}")
    print(f"  tools ever dropped     : {sum(p['dropped_tools'] for p in planning)}")
    print(f"  repeated identical calls: {repeats or 'none'}")
    print(f"  planner errors         : "
          f"{[p['error'] for p in planning if p['error']] or 'none'}")

    payload = run.to_dict()
    payload["planning"] = planning
    payload["repeats"] = repeats
    out = Path(__file__).parent.parent / "design" / "evidence" / f"longhorizon_run{args.run}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, indent=2, default=str))
    print(f"evidence: {out}")


if __name__ == "__main__":
    main()
